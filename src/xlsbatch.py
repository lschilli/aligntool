import fnmatch
import os
import shlex
import util
import logging
import tgt
import argparse

import aligntool
from openpyxl import Workbook, load_workbook, comments
from collections import defaultdict, namedtuple

#todo: catch exceptions and print source column meta info
#todo: reset onset cell range


class ArgParseException(Exception):
    def __init__(self, message, helptext):
        self.message = message
        self.helptext = helptext

    def __str__(self):
        return repr(self.message)


class ArgParserWrapper(argparse.ArgumentParser):

    def _print_message(self, message, file=None):
        pass

    def exit(self, status=0, message=None):
        raise ArgParseException(message, self.format_help())


def estimate_col_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 20), len(str(cell.value))*0.7))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value


class WavImporter:

    def __init__(self, sub_cmd_parser=None):
        if sub_cmd_parser:
            cmd_parser = sub_cmd_parser.add_parser('wavlist', help='import a list of wav files from a directory tree',
                                                   formatter_class=argparse.ArgumentDefaultsHelpFormatter)
            cmd_parser.set_defaults(cmd=self.run)
            cmd_parser.add_argument('-x', dest='xlsxfile', metavar='<xlsfile>', action='store', type=str,
                                    required=True, help='xlsxfile')
            cmd_parser.add_argument('-d', dest='directory', metavar='<d>', action='store', type=str,
                                    required=True, help='directory <d> to scan for wav files')

    def run(self, directory, xlsxfile):
        if os.path.exists(xlsxfile):
            wb = load_workbook(filename=xlsxfile)
            if 'batch' in wb.get_sheet_names():
                ws = wb.get_sheet_by_name('batch')
            else:
                ws = wb.create_sheet('batch')
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "batch"
        i = 0
        for head in ("Wavefile", "TextGrid"):
            ws.cell(row=1, column=i+1).value = head
            i += 1
        for head in ("segmentBeeps", "addTier", "segmentSpeech", "alignMAUS", "extractOnOffsets"):
            ws.cell(row=1, column=i + 1).value = head
            try:
                hcp = ArgParserWrapper()
                if head == "extractOnOffsets":
                    oe = OnOffsetExtractor()
                    oe.get_option_parser().parse_args(shlex.split(head))
                else:
                    aligntool.parse_arguments(shlex.split(head), hcp)
            except ArgParseException as e:
                comment = comments.Comment(e.helptext, 'aligntool')
                comment._width = '400pt'
                ws.cell(row=1, column=i + 1).comment = comment
            i += 1
        row = 2
        uniquetextgrids = set()
        xlsxprefix, _ = os.path.splitext(xlsxfile)
        workdir = xlsxprefix + ".tg"
        for root, dirs, files in os.walk(directory, topdown=True):
            for f in fnmatch.filter(files, "*.[wW][aA][vV]"):
                fname = os.path.join(root, f)
                froot, ext = os.path.splitext(f)
                fstr = ""
                fnum = 0
                while os.path.join(workdir, froot + fstr + ".TextGrid") in uniquetextgrids:
                    fnum += 1
                    fstr = str(fnum)
                ws.cell(row=row, column=1).value = fname
                gridfile = os.path.join(workdir, froot + fstr + ".TextGrid")
                ws.cell(row=row, column=2).value = gridfile
                uniquetextgrids.add(gridfile)
                row += 1
        for i in range(row, ws.max_row):
            for j in range(1, 3):
                ws.cell(row=i, column=j).value = None
        estimate_col_width(ws)
        wb.save(xlsxfile)


# Performs a batch run based on parameters in the files sheet
class BatchRunner:
    def __init__(self, sub_cmd_parser=None):
        if sub_cmd_parser:
            cmd_parser = sub_cmd_parser.add_parser('run', help='start a batch run',
                                                   formatter_class=argparse.ArgumentDefaultsHelpFormatter)
            cmd_parser.set_defaults(cmd=self.run)
            cmd_parser.add_argument('-x', dest='xlsxfile', metavar='<xlsfile>', action='store', type=str,
                                    required=True, help='xlsxfile with batch sheet')
            cmd_parser.add_argument('-c', dest='batchcmd', metavar='<cmd>', action='store', type=str,
                                    required=True, help='command column to run')

    def run(self, xlsxfile, batchcmd):
        wb = load_workbook(filename=xlsxfile, read_only=True)
        ws = wb['batch']
        coldict = {}
        try:
            for i, row in enumerate(ws.rows):
                if i == 0:
                    coldict = {cell.value: i for i, cell in enumerate(row)}
                    continue
                wavfile = row[coldict['Wavefile']].value
                textgrid = row[coldict['TextGrid']].value
                if wavfile is None:
                    logging.warning("Ignoring row %s: Wavefile column empty" % (i+1))
                    continue
                if textgrid is None:
                    logging.warning("Ignoring row %s: TextGrid column empty" % (i+1))
                    continue
                cmdparams = row[coldict[batchcmd]].value
                cmdparams = "" if cmdparams is None else cmdparams
                logging.debug(wavfile, textgrid, cmdparams)
                tgopts = []
                if os.path.exists(textgrid):
                    tgopts = ['-i', textgrid]
                else:
                    path = os.path.dirname(textgrid)
                    os.makedirs(name=path, exist_ok=True)
                cmd = [batchcmd] + tgopts + ['-o', textgrid, '-w', wavfile] + shlex.split(cmdparams)
                logging.info("Row %s: running: %s" % (i+1, " ".join(cmd)))
                apw = ArgParserWrapper("")
                args = aligntool.parse_arguments(cmd, apw)
                args.cmd(**util.extract_args(args))
        except ArgParseException as e:
            logging.error("Batch processing failed in row %s:%s" % (i+1, str(e)))
        except Exception as e:
            logging.error("Batch processing failed in row %s: %s" % (i+1, str(e)))


# Imports generated Textgrids, e.g. after each command run
class TextGridBulkImporter:

    def __init__(self, sub_cmd_parser=None):
        if sub_cmd_parser:
            cmd_parser = sub_cmd_parser.add_parser('tgimport', help='import textgrid listed in the batch sheet',
                                                   formatter_class=argparse.ArgumentDefaultsHelpFormatter)
            cmd_parser.set_defaults(cmd=self.run)
            cmd_parser.add_argument('-x', dest='xlsxfile', metavar='<xlsfile>', action='store', type=str,
                                    required=True, help='xlsxfile with batch sheet')

    def run(self, xlsxfile):
        logging.info("Opening %s" % xlsxfile)
        wb = load_workbook(filename=xlsxfile)
        ws = wb['batch']
        if 'segments' in wb.get_sheet_names():
            segws = wb.get_sheet_by_name('segments')
        else:
            segws = wb.create_sheet('segments')

        header = ["TextGrid", "TierName", "StartTime", "EndTime", "Text"]
        for i, head in enumerate(header):
            segws.cell(row=1, column=i+1).value = head

        coldict = {}
        rowcnt = 1
        for i, row in enumerate(ws.rows):
            if i == 0:
                coldict = {cell.value: i for i, cell in enumerate(row)}
                continue
            rowcnt += 1
            textgridfilename = row[coldict['TextGrid']].value
            if textgridfilename is None:
                continue
            logging.info("Reading %s" % textgridfilename)
            tg = tgt.io.read_textgrid(textgridfilename)
            rowvalues = [textgridfilename, "<range>", float(tg.start_time), float(tg.end_time), ""]
            for j, val in enumerate(rowvalues):
                segws.cell(row=rowcnt, column=j + 1).value = val
            for tier in tg:
                for iv in tier:
                    rowcnt += 1
                    rowvalues = [textgridfilename, tier.name, float(iv.start_time), float(iv.end_time), iv.text]
                    for j, val in enumerate(rowvalues):
                        segws.cell(row=rowcnt, column=j+1).value = val
        for i in range(rowcnt+1, segws.max_row+1):
            for j in range(1, len(header)+1):
                segws.cell(row=i, column=j).value = None
        estimate_col_width(segws)
        wb.save(xlsxfile)


# Generates Textgrids, e.g. after xls edits
class TextGridBulkExporter:
    def __init__(self, sub_cmd_parser=None):
        if sub_cmd_parser:
            cmd_parser = sub_cmd_parser.add_parser('tgexport', help='import textgrid listed in the batch sheet',
                                                   formatter_class=argparse.ArgumentDefaultsHelpFormatter)
            cmd_parser.set_defaults(cmd=self.run)
            cmd_parser.add_argument('-x', dest='xlsxfile', metavar='<xlsfile>', action='store', type=str,
                                    required=True, help='xlsxfile with segments sheet')

    def set_range(self, tg):
        if hasattr(tg, 'range_start_time'):
            tiers = tg.get_tier_names()
            if len(tiers) > 0:
                tier = tg.get_tier_by_name(tiers[0])
                tier.start_time = tg.range_start_time
                tier.end_time = tg.range_end_time

    def run(self, xlsxfile):
        logging.info("Opening %s" % xlsxfile)
        wb = load_workbook(filename=xlsxfile)
        assert 'segments' in wb.get_sheet_names(), "sheet segments, required for export, does not exist in %s" % xlsxfile
        segws = wb.get_sheet_by_name('segments')
        tgdict = defaultdict(tgt.TextGrid)
        for seg_row_num, seg_row in enumerate(segws.rows):
            if seg_row_num == 0:
                segcoldict = {cell.value: i for i, cell in enumerate(seg_row)}
                continue
            textgrid = seg_row[segcoldict['TextGrid']].value
            tiername = seg_row[segcoldict['TierName']].value
            tbegin = seg_row[segcoldict['StartTime']].value
            tend = seg_row[segcoldict['EndTime']].value
            text = seg_row[segcoldict['Text']].value
            if None in (textgrid, tiername, tbegin, tend):
                continue
            tg = tgdict[textgrid]
            if tiername == "<range>":
                tg.range_end_time = tend
                tg.range_start_time = tbegin
                self.set_range(tg)
                continue
            if not text:
                continue
            if not tg.has_tier(tiername):
                tier = tgt.IntervalTier(name=tiername)
                tg.add_tier(tier)
            else:
                tier = tg.get_tier_by_name(tiername)
            iv = tgt.Annotation(tbegin, tend, text)
            tier.add_annotation(iv)
            self.set_range(tg)
        for filename, tg in tgdict.items():
            logging.info("Writing %s" % filename)
            path = os.path.dirname(filename)
            os.makedirs(name=path, exist_ok=True)
            tgt.io.write_to_file(tg, filename)


class OnOffsetExtractor:

    def __init__(self, sub_cmd_parser=None):
        self.rowpos = 1
        self.header = ("TextGrid", "Error", "ErrorMsg", "Transcription", "AlignOnset", "AlignOffset", "WordCount",
                       "NumWords", "OnsetPhoneme", "OffsetPhoneme", "PresegBegin")
        self.ColHead = namedtuple("ColHead", self.header)
        self.colids = self.ColHead(*range(1, len(self.header) + 1))
        if sub_cmd_parser:
            cmd_parser = sub_cmd_parser.add_parser('extractOnOffsets', help='find onsets and offsets within a filter '
                                                                            'interval and save it to a OnsetOffset '
                                                                            'Excel sheet',
                                                   formatter_class=argparse.ArgumentDefaultsHelpFormatter)
            cmd_parser.set_defaults(cmd=self.run)
            cmd_parser.add_argument('-x', dest='xlsxfile', metavar='<xlsfile>', action='store', type=str,
                                    required=True, help='xlsxfile with batch sheet')

    def get_option_parser(self):
        cmd_parser = ArgParserWrapper("extractOnOffsets")
        cmd_parser.add_argument('-f', "--filter-tier", dest='filtertiername', metavar='<tier>',
                                action='store',
                                default="seg.beep",
                                help='pre-segmentation tier')
        return cmd_parser

    def run(self, xlsxfile):
        logging.info("Opening %s" % xlsxfile)
        wb = load_workbook(filename=xlsxfile)
        ws = wb['batch']
        if 'on_offsets' in wb.get_sheet_names():
            statws = wb.get_sheet_by_name('on_offsets')
        else:
            statws = wb.create_sheet('on_offsets')
        for i, head in enumerate(self.header):
            statws.cell(row=1, column=i + 1).value = head
        optionparser = self.get_option_parser()
        coldict = {}
        try:
            for i, row in enumerate(ws.rows):
                if i == 0:
                    coldict = {cell.value: i for i, cell in enumerate(row)}
                    continue
                textgridfilename = row[coldict['TextGrid']].value
                if textgridfilename is None:
                    continue
                cmdparams = row[coldict['extractOnOffsets']].value
                if cmdparams is None:
                    cmdparams = ""
                options = optionparser.parse_args(shlex.split(cmdparams))
                self.on_offsets_from_tg(textgridfilename, options.filtertiername, statws)
            for i in range(self.rowpos+1, statws.max_row+1):
                for j in range(1, len(self.header)+1):
                    statws.cell(row=i, column=j).value = None
            wb.save(xlsxfile)
        except Exception as e:
            logging.error("Batch processing failed in row %s: %s" % (i+1, str(e)))
            raise

    def on_offsets_from_tg(self, textgridfilename, filtertiername, ws):
        logging.info("Reading %s" % textgridfilename)
        try:
            tg = tgt.io.read_textgrid(textgridfilename)
            align_tier = tg.get_tier_by_name("maus.ort")
            pho_tier = tg.get_tier_by_name("maus.pho")
            seg_tier = tg.get_tier_by_name(filtertiername)
            for seg_iv in [s for s in seg_tier if s.text == "speech"]:
                word_cnt = 1
                align_ivs = align_tier.get_annotations_between_timepoints(seg_iv.start_time, seg_iv.end_time,
                                                                          left_overlap=True, right_overlap=True)
                for aligniv in align_ivs:
                    self.rowpos += 1
                    ws.cell(row=self.rowpos, column=self.colids.TextGrid).value = tg.filename
                    ws.cell(row=self.rowpos, column=self.colids.Transcription).value = aligniv.text
                    ws.cell(row=self.rowpos, column=self.colids.NumWords).value = len(align_ivs)
                    ws.cell(row=self.rowpos, column=self.colids.Error).value = 0
                    ws.cell(row=self.rowpos, column=self.colids.ErrorMsg).value = ""
                    ws.cell(row=self.rowpos, column=self.colids.PresegBegin).value = float(seg_iv.start_time)
                    ws.cell(row=self.rowpos, column=self.colids.AlignOnset).value = float(aligniv.start_time)
                    ws.cell(row=self.rowpos, column=self.colids.AlignOffset).value = float(aligniv.end_time)
                    pho_ivs = pho_tier.get_annotations_between_timepoints(aligniv.start_time, aligniv.end_time,
                                                                          left_overlap=True, right_overlap=True)
                    pho_ivs_cleaned = [x.text for x in pho_ivs if x.text not in ("?", "<p:>", "<usb>")]
                    if pho_ivs_cleaned:
                        ws.cell(row=self.rowpos, column=self.colids.OnsetPhoneme).value = pho_ivs_cleaned[0]
                        ws.cell(row=self.rowpos, column=self.colids.OffsetPhoneme).value = pho_ivs_cleaned[-1]
                    else:
                        msg = "Phonetic alignment for %s seems empty after filtering: %s" % (aligniv, pho_ivs_cleaned)
                        logging.warning(msg)
                        ws.cell(row=self.rowpos, column=self.colids.OnsetPhoneme).value = ""
                        ws.cell(row=self.rowpos, column=self.colids.OffsetPhoneme).value = ""
                        ws.cell(row=self.rowpos, column=self.colids.Error).value = 1
                        ws.cell(row=self.rowpos, column=self.colids.ErrorMsg).value = msg
                    ws.cell(row=self.rowpos, column=self.colids.WordCount).value = word_cnt
                    word_cnt += 1
        except Exception as e:
            logging.error("On/Offset extraction failed for %s: %s" % (textgridfilename, e))
            self.rowpos += 1
            ws.cell(row=self.rowpos, column=self.colids.TextGrid).value = textgridfilename
            ws.cell(row=self.rowpos, column=self.colids.Error).value = 1
            ws.cell(row=self.rowpos, column=self.colids.ErrorMsg).value = str(e)


def setup(sub_cmd_parser):
    importers = [WavImporter, BatchRunner, TextGridBulkImporter, TextGridBulkExporter, OnOffsetExtractor]
    for imp in importers:
        imp(sub_cmd_parser)
