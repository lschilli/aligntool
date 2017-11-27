import inspect
import json
import logging
import shlex
import subprocess
import openpyxl
import tgt
import wave


def get_wav_duration(filename):
    """
    :param filename: wavfile
    :return: length in seconds
    """
    with wave.open(filename, 'r') as f:
        samplerate = f.getframerate()
        duration = f.getnframes() / float(samplerate)
        return duration, samplerate


def tier_to_str(tier):
    return " ".join(["%s[%.2f..%.2f]" % (x.text, x.start_time, x.end_time) for x in tier])


def extract_optional_args(args):
    names, varargs, varkw, defaults = inspect.getargspec(args.cmd)
    keys = names[-len(defaults):]
    d = vars(args)
    return dict((k, d[k]) for k in keys if k in d and d[k] is not None)


def extract_args(args):
    names, varargs, varkw, defaults = inspect.getargspec(args.cmd)
    d = vars(args)
    return dict((k, d[k]) for k in names if k in d)


def call_check(cmd, return_stdout=False, stdindata=None, return_stderr=False):
    if isinstance(cmd, list):
        logmsg = "exec[]: " + " ".join(cmd)
        procname = cmd[0]
        shell = False
    else:
        logmsg = "exec: " + cmd
        procname = shlex.split(cmd)[0]
        shell = True

    pipein, pipeout, pipeerr = None, None, None
    if return_stdout:
        pipeout = subprocess.PIPE

    if return_stderr:
        pipeerr = subprocess.PIPE

    if stdindata is not None:
        pipein = subprocess.PIPE

    logging.info(logmsg)
    proc = subprocess.Popen(cmd, shell=shell, stdin=pipein, stdout=pipeout, stderr=pipeerr)
    if stdindata is not None:
        proc.stdin.write(stdindata)
        proc.stdin.close()
    output = None
    output_raw = proc.communicate()
    if return_stdout:
        output = output_raw[0]
    if return_stderr:
        output = output_raw[1]
    try:
        exitcode = proc.wait()
    except KeyboardInterrupt:
        exitcode = proc.wait()
    if exitcode != 0:
        errmsg = "%s exited with non-zero status" % procname
        logging.error(errmsg)
        raise subprocess.CalledProcessError(exitcode, cmd)
    return output


def init_textgrid(infile, duration, *tiers):
    tg = tgt.TextGrid()
    if infile is not None:
        logging.info("reading TextGrid %s" % infile)
        tg = tgt.io.read_textgrid(infile)

    result = [tg]
    for tier in tiers:
        if tg.has_tier(tier):
            logging.info("overwriting tier %s" % tier)
            tg.delete_tier(tier)
        tier = tgt.IntervalTier(name=tier, start_time=0, end_time=duration)
        result.append(tier)
    return tuple(result)


def get_sheet(infile, sheetname):
    wb = openpyxl.load_workbook(filename=infile, read_only=True)
    sheet = wb[sheetname]
    cols = sheet.max_column
    headers = dict((sheet.cell(row=1, column=i).value, i - 1) for i in range(1, cols + 1))
    return sheet, headers


def filter_data_row(values, prevvalues, datafilter):
    evalvars = {"row": values, "prevrow": prevvalues}
    try:
        ret = eval(datafilter, evalvars, evalvars)
    except:
        logging.error("Exception parsing/evaluating datafilter (%s)" % datafilter)
        raise
    assert ret in (True, False), "Filter function must return either true or false but not %s" % ret
    return ret


class MetaData:
    def __init__(self, ref=None, valid=True, orig="", edited=False, ignore=False, shortened=False):
        self.ref = ref
        self.valid = valid
        self.orig = orig
        self.edited = edited
        self.ignore = ignore
        self.shortened = shortened

    def __str__(self):
        return str(self.__dict__)

    def to_json(self):
        return json.dumps(self.__dict__, sort_keys=True)

    @classmethod
    def from_json(cls, jsonstr):
        meta = cls(**json.loads(jsonstr))
        return meta
